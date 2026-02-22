from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from datetime import datetime

# Color palette
NAVY = "1B3A6B"
TEAL = "0D7377"
WHITE = "FFFFFF"
LIGHT_BLUE = "E8F0F7"
LIGHT_TEAL = "E6F4F1"
LIGHT_GRAY = "F3F4F6"
GREEN_FILL = "D1FAE5"
GREEN_TEXT = "047857"
AMBER_FILL = "FEF3C7"
AMBER_TEXT = "D97706"
RED_FILL = "FEE2E2"
RED_TEXT = "DC2626"
YELLOW_FILL = "FEF9C3"
YELLOW_TEXT = "92400E"
GRAY = "6B7280"

thin_side = Side(style="thin", color="D1D5DB")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_font(bold=False, color=None, size=10, italic=False) -> Font:
    return Font(
        name="Calibri",
        bold=bold,
        color=color or "000000",
        size=size,
        italic=italic,
    )

def write_excel(result: dict, output_path: str):
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Income Statement"

    metadata = result.get("extraction_metadata", {})
    years = result.get("years_detected", [])
    line_items = result.get("line_items", [])
    currency = metadata.get("currency", "?")
    unit = metadata.get("unit", "?")
    warnings = metadata.get("warnings", [])
    validation_status = metadata.get("validation_status", "UNKNOWN")

    # ─── HEADER BANNER ────────────────────────────────────────────────────
    ws_main.merge_cells("A1:A2")
    ws_main.merge_cells(f"B1:{get_column_letter(3 + len(years))}2")

    banner_cell = ws_main["A1"]
    banner_cell.value = "INCOME STATEMENT EXTRACTION"
    banner_cell.fill = make_fill(NAVY)
    banner_cell.font = make_font(bold=True, color=WHITE, size=14)
    banner_cell.alignment = Alignment(horizontal="center", vertical="center")

    info_cell = ws_main["B1"]
    info_cell.value = (
        f"Currency: {currency}  |  Unit: {unit}  |  "
        f"Source: {metadata.get('source_context_notes', 'Annual Report')}  |  "
        f"Validation: {validation_status}  |  "
        f"Extracted: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    )
    info_cell.fill = make_fill(NAVY)
    info_cell.font = make_font(color="CBD5E1", size=9)
    info_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Warnings row
    row_offset = 3
    if warnings:
        ws_main.merge_cells(f"A{row_offset}:{get_column_letter(3 + len(years))}{row_offset}")
        warn_cell = ws_main[f"A{row_offset}"]
        warn_cell.value = "⚠ VALIDATION WARNINGS: " + "  |  ".join(warnings)
        warn_cell.fill = make_fill(AMBER_FILL)
        warn_cell.font = make_font(bold=True, color=AMBER_TEXT, size=9)
        warn_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_main.row_dimensions[row_offset].height = 30
        row_offset += 1

    # ─── COLUMN HEADERS ────────────────────────────────────────────────────
    col_headers = ["Line Item (Canonical)", "Source Label"] + years + ["Confidence", "Source Pages", "Notes"]
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws_main.cell(row=row_offset, column=col_idx, value=header)
        cell.fill = make_fill(NAVY)
        cell.font = make_font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_main.row_dimensions[row_offset].height = 28
    row_offset += 1

    # ─── DATA ROWS ─────────────────────────────────────────────────────────
    # Group line items into sections
    SECTIONS = {
        "Revenue": ["Revenue"],
        "Cost & Gross Profit": ["COGS", "Gross Profit"],
        "Operating Expenses": ["R&D Expenses", "SG&A Expenses", "Other Operating Expenses", "Total Operating Expenses"],
        "Operating Results": ["Operating Income", "EBITDA", "Depreciation & Amortization"],
        "Below-the-Line": ["Interest Expense", "Interest Income", "Other Income/Expense"],
        "Pre-Tax & Tax": ["Income Before Tax", "Income Tax Expense"],
        "Bottom Line": ["Net Income"],
        "Per Share": ["Basic EPS", "Diluted EPS", "Basic Shares Outstanding", "Diluted Shares Outstanding"],
    }

    # Build lookup
    li_map = {li["canonical_name"]: li for li in line_items}

    # Add all line items not in sections
    all_section_items = [item for items in SECTIONS.values() for item in items]
    extra_items = [li["canonical_name"] for li in line_items if li["canonical_name"] not in all_section_items]

    section_fill = make_fill(LIGHT_BLUE)
    alt_fill = make_fill(LIGHT_GRAY)

    row_num = row_offset
    alt = False

    for section_name, items in SECTIONS.items():
        # Section header row
        ws_main.merge_cells(f"A{row_num}:{get_column_letter(len(col_headers))}{row_num}")
        sec_cell = ws_main[f"A{row_num}"]
        sec_cell.value = section_name.upper()
        sec_cell.fill = make_fill(TEAL)
        sec_cell.font = make_font(bold=True, color=WHITE, size=9)
        sec_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sec_cell.border = thin_border
        ws_main.row_dimensions[row_num].height = 18
        row_num += 1

        for item_name in items:
            li = li_map.get(item_name)
            row_fill = alt_fill if alt else make_fill(WHITE)
            alt = not alt

            confidence = li["confidence"] if li else None
            values = li.get("values", {}) if li else {}
            source_label = li.get("source_label") if li else None
            notes = li.get("notes") if li else None
            source_pages = li.get("source_pages", []) if li else []

            # Determine if any value exists
            has_value = li and any(v is not None for v in values.values())

            # Confidence fill
            if not li or not has_value:
                conf_fill = make_fill(YELLOW_FILL)
                conf_color = YELLOW_TEXT
            elif confidence == "HIGH":
                conf_fill = make_fill(GREEN_FILL)
                conf_color = GREEN_TEXT
            elif confidence == "MEDIUM":
                conf_fill = make_fill(AMBER_FILL)
                conf_color = AMBER_TEXT
            else:
                conf_fill = make_fill(RED_FILL)
                conf_color = RED_TEXT

            # Col A: Canonical name
            a_cell = ws_main.cell(row=row_num, column=1, value=item_name)
            a_cell.fill = make_fill(LIGHT_TEAL)
            a_cell.font = make_font(bold=True, color=NAVY, size=10)
            a_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            a_cell.border = thin_border

            # Col B: Source label
            b_cell = ws_main.cell(row=row_num, column=2, value=source_label or "—")
            b_cell.fill = row_fill
            b_cell.font = make_font(color=GRAY, size=9, italic=True)
            b_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            b_cell.border = thin_border

            # Year value columns
            for col_offset, year in enumerate(years):
                col_num = 3 + col_offset
                val = values.get(year)

                if val is None:
                    display = "Not Reported"
                    val_fill = make_fill(YELLOW_FILL)
                    val_font = make_font(color=YELLOW_TEXT, size=9, italic=True)
                    num_format = "@"  # Text
                else:
                    display = val
                    val_fill = row_fill
                    val_font = make_font(color="111827", size=10)
                    num_format = '#,##0.00'

                v_cell = ws_main.cell(row=row_num, column=col_num, value=display)
                v_cell.fill = val_fill
                v_cell.font = val_font
                v_cell.alignment = Alignment(horizontal="right", vertical="center")
                v_cell.border = thin_border
                if isinstance(display, (int, float)):
                    v_cell.number_format = num_format

            # Confidence
            conf_col = 3 + len(years)
            c_cell = ws_main.cell(row=row_num, column=conf_col, value=confidence or "N/A")
            c_cell.fill = conf_fill
            c_cell.font = make_font(bold=True, color=conf_color, size=9)
            c_cell.alignment = Alignment(horizontal="center", vertical="center")
            c_cell.border = thin_border

            # Source pages
            pages_str = f"p.{','.join(str(p) for p in source_pages)}" if source_pages else "—"
            sp_cell = ws_main.cell(row=row_num, column=conf_col + 1, value=pages_str)
            sp_cell.fill = row_fill
            sp_cell.font = make_font(color=GRAY, size=9)
            sp_cell.alignment = Alignment(horizontal="center", vertical="center")
            sp_cell.border = thin_border

            # Notes
            n_cell = ws_main.cell(row=row_num, column=conf_col + 2, value=notes or "")
            n_cell.fill = row_fill
            n_cell.font = make_font(color=GRAY, size=9, italic=bool(notes))
            n_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            n_cell.border = thin_border

            ws_main.row_dimensions[row_num].height = 18
            row_num += 1

    # Extra items not in schema (if any)
    if extra_items:
        ws_main.merge_cells(f"A{row_num}:{get_column_letter(len(col_headers))}{row_num}")
        sec_cell = ws_main[f"A{row_num}"]
        sec_cell.value = "ADDITIONAL ITEMS DETECTED"
        sec_cell.fill = make_fill("7C3AED")
        sec_cell.font = make_font(bold=True, color=WHITE, size=9)
        sec_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sec_cell.border = thin_border
        row_num += 1

        for item_name in extra_items:
            li = li_map.get(item_name)
            if not li:
                continue
            values = li.get("values", {})
            a_cell = ws_main.cell(row=row_num, column=1, value=item_name)
            a_cell.fill = make_fill("F3E8FF")
            a_cell.font = make_font(bold=True, color="7C3AED", size=10)
            a_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            a_cell.border = thin_border

            b_cell = ws_main.cell(row=row_num, column=2, value=li.get("source_label") or "—")
            b_cell.fill = make_fill(WHITE)
            b_cell.font = make_font(color=GRAY, size=9, italic=True)
            b_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            b_cell.border = thin_border

            for col_offset, year in enumerate(years):
                col_num = 3 + col_offset
                val = values.get(year)
                v_cell = ws_main.cell(row=row_num, column=col_num, value=val if val is not None else "Not Reported")
                v_cell.fill = make_fill(WHITE)
                v_cell.border = thin_border
                v_cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, (int, float)):
                    v_cell.number_format = '#,##0.00'

            row_num += 1

    # ─── LEGEND ROW ────────────────────────────────────────────────────────
    row_num += 1
    legend_items = [
        (GREEN_FILL, GREEN_TEXT, "HIGH confidence"),
        (AMBER_FILL, AMBER_TEXT, "MEDIUM confidence"),
        (RED_FILL, RED_TEXT, "LOW confidence"),
        (YELLOW_FILL, YELLOW_TEXT, "Not Reported"),
    ]
    ws_main.cell(row=row_num, column=1, value="LEGEND:").font = make_font(bold=True, color=NAVY, size=9)
    for i, (fill, text_color, label) in enumerate(legend_items, start=2):
        cell = ws_main.cell(row=row_num, column=i, value=label)
        cell.fill = make_fill(fill)
        cell.font = make_font(bold=True, color=text_color, size=9)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # ─── COLUMN WIDTHS ─────────────────────────────────────────────────────
    ws_main.column_dimensions["A"].width = 30
    ws_main.column_dimensions["B"].width = 28
    for i, year in enumerate(years):
        ws_main.column_dimensions[get_column_letter(3 + i)].width = 16
    conf_col_letter = get_column_letter(3 + len(years))
    ws_main.column_dimensions[conf_col_letter].width = 12
    ws_main.column_dimensions[get_column_letter(3 + len(years) + 1)].width = 12
    ws_main.column_dimensions[get_column_letter(3 + len(years) + 2)].width = 40

    # Freeze panes
    ws_main.freeze_panes = f"C{row_offset}"

    # ─── METADATA TAB ──────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Extraction Metadata")
    meta_rows = [
        ("Field", "Value"),
        ("Source File", metadata.get("source_file", "uploaded_document.pdf")),
        ("Extraction Timestamp", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Currency", metadata.get("currency", "?")),
        ("Unit", metadata.get("unit", "?")),
        ("Fiscal Year End", metadata.get("fiscal_year_end", "Unknown")),
        ("Years Detected", ", ".join(result.get("years_detected", []))),
        ("Source Pages", str(metadata.get("source_pages", []))),
        ("Total PDF Pages", str(metadata.get("total_pdf_pages", "?"))),
        ("OCR Source", str(metadata.get("ocr_source", False))),
        ("Validation Status", metadata.get("validation_status", "UNKNOWN")),
        ("Validation Warnings", "\n".join(metadata.get("warnings", [])) or "None"),
        ("Extraction Model", "claude-sonnet-4-6"),
        ("Schema Version", "v1.0"),
        ("Context Notes", metadata.get("source_context_notes", "")),
    ]
    for row_idx, (field, value) in enumerate(meta_rows, start=1):
        f_cell = ws_meta.cell(row=row_idx, column=1, value=field)
        v_cell = ws_meta.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            f_cell.fill = make_fill(NAVY)
            f_cell.font = make_font(bold=True, color=WHITE)
            v_cell.fill = make_fill(NAVY)
            v_cell.font = make_font(bold=True, color=WHITE)
        else:
            f_cell.fill = make_fill(LIGHT_TEAL)
            f_cell.font = make_font(bold=True, color=NAVY)
            v_cell.fill = make_fill(WHITE)
            v_cell.font = make_font()
        f_cell.border = thin_border
        v_cell.border = thin_border
        f_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        v_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws_meta.row_dimensions[row_idx].height = 22

    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 60

    # ─── INSTRUCTIONS TAB ──────────────────────────────────────────────────
    ws_help = wb.create_sheet("How to Read This")
    help_rows = [
        ("HOW TO READ THIS WORKBOOK", "", ""),
        ("Sheet", "Description", ""),
        ("Income Statement", "Main extraction output. All canonical line items with values per year.", ""),
        ("Extraction Metadata", "Full audit trail: source file, model used, pages processed, validation results.", ""),
        ("How to Read This", "This guide.", ""),
        ("", "", ""),
        ("CONFIDENCE LEVELS", "", ""),
        ("HIGH", "Value found with high certainty. Exact or alias match.", "Green"),
        ("MEDIUM", "Fuzzy match or slight ambiguity in source label.", "Amber"),
        ("LOW", "Uncertain extraction. Verify against source document.", "Red"),
        ("Not Reported", "Line item not found in the document.", "Yellow"),
        ("", "", ""),
        ("IMPORTANT NOTES", "", ""),
        ("1.", "All values are in the unit shown in the header (e.g., millions USD).", ""),
        ("2.", "Source Label column shows the exact text from the document.", ""),
        ("3.", "Source Pages column shows which PDF pages contained each item.", ""),
        ("4.", "Validation Warnings (if any) are shown in the amber banner at the top of Income Statement tab.", ""),
        ("5.", "This tool extracts only what is present in the document. It does NOT estimate missing values.", ""),
    ]
    for row_idx, (c1, c2, c3) in enumerate(help_rows, start=1):
        for col_idx, val in enumerate([c1, c2, c3], start=1):
            cell = ws_help.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.fill = make_fill(NAVY)
                cell.font = make_font(bold=True, color=WHITE, size=12)
            elif val in ("Sheet", "CONFIDENCE LEVELS", "IMPORTANT NOTES"):
                cell.fill = make_fill(TEAL)
                cell.font = make_font(bold=True, color=WHITE)
            elif c1 in ("HIGH", "MEDIUM", "LOW", "Not Reported") and col_idx == 1:
                fills = {"HIGH": GREEN_FILL, "MEDIUM": AMBER_FILL, "LOW": RED_FILL, "Not Reported": YELLOW_FILL}
                colors = {"HIGH": GREEN_TEXT, "MEDIUM": AMBER_TEXT, "LOW": RED_TEXT, "Not Reported": YELLOW_TEXT}
                cell.fill = make_fill(fills.get(val, WHITE))
                cell.font = make_font(bold=True, color=colors.get(val, "000000"))
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

    ws_help.column_dimensions["A"].width = 20
    ws_help.column_dimensions["B"].width = 70
    ws_help.column_dimensions["C"].width = 15

    wb.save(output_path)
